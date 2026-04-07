#!/usr/bin/env python3
"""api_server.py — EIS Investor Collector backend."""
import os
import json
import math
import io
from pathlib import Path
from datetime import datetime, timedelta
from contextlib import asynccontextmanager
from typing import Optional

import psycopg2
import psycopg2.extras

from fastapi import FastAPI, Query, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers


def get_db():
    database_url = os.environ.get("DATABASE_URL", "")
    if database_url:
        conn = psycopg2.connect(database_url, cursor_factory=psycopg2.extras.RealDictCursor)
    else:
        # Fallback: local SQLite-style is not available with psycopg2.
        # Raise a clear error if DATABASE_URL is not set.
        raise RuntimeError(
            "DATABASE_URL environment variable is not set. "
            "Please configure a PostgreSQL database."
        )
    return conn


def init_db(db):
    cur = db.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS investors (
            id SERIAL PRIMARY KEY,
            name TEXT NOT NULL,
            role TEXT,
            company TEXT,
            eis_company TEXT,
            sector TEXT,
            amount TEXT,
            source_url TEXT,
            source_type TEXT,
            source_name TEXT,
            context_quote TEXT,
            linkedin_url TEXT,
            date_found TEXT,
            created_at TIMESTAMP DEFAULT NOW(),
            bio TEXT
        )
    """)
    # Add bio column for databases created before this schema update
    cur.execute("ALTER TABLE investors ADD COLUMN IF NOT EXISTS bio TEXT")
    cur.execute("""
        CREATE TABLE IF NOT EXISTS export_log (
            id SERIAL PRIMARY KEY,
            exported_at TIMESTAMP DEFAULT NOW(),
            investor_count INTEGER DEFAULT 0,
            export_type TEXT DEFAULT 'full'
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS email_log (
            id SERIAL PRIMARY KEY,
            emailed_at TIMESTAMP DEFAULT NOW(),
            investor_count INTEGER DEFAULT 0
        )
    """)
    cur.execute("CREATE INDEX IF NOT EXISTS idx_name ON investors(name)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_date ON investors(date_found)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_sector ON investors(sector)")
    cur.execute("CREATE INDEX IF NOT EXISTS idx_created ON investors(created_at)")
    db.commit()
    cur.close()


def seed_db(db):
    cur = db.cursor()
    cur.execute("SELECT COUNT(*) as c FROM investors")
    row = cur.fetchone()
    count = row["c"]
    cur.close()

    if count > 0:
        return

    # Load seed data from JSON file bundled with the app
    seed_file = Path(__file__).parent / "seed_data.json"
    if seed_file.exists():
        with open(seed_file) as f:
            seed_data = json.load(f)
        cur = db.cursor()
        for inv in seed_data:
            cur.execute("""
                INSERT INTO investors (name, role, company, eis_company, sector, amount,
                source_url, source_type, source_name, context_quote, linkedin_url, date_found)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                inv.get("name"), inv.get("role"), inv.get("company"), inv.get("eis_company"),
                inv.get("sector"), inv.get("amount"), inv.get("source_url"), inv.get("source_type"),
                inv.get("source_name"), inv.get("context_quote"), inv.get("linkedin_url"), inv.get("date_found")
            ))
        db.commit()
        cur.close()
        print(f"[seed] Loaded {len(seed_data)} investors from seed_data.json")
        return

    print("[seed] No seed_data.json found. Database starts empty.")


# --- Scan History Tracking ---

def _init_scan_history():
    """Create table to track when each scan type last ran."""
    db = get_db()
    cur = db.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS scan_history (
            id SERIAL PRIMARY KEY,
            scan_type TEXT NOT NULL,
            started_at TIMESTAMP NOT NULL,
            finished_at TIMESTAMP,
            result TEXT
        )
    """)
    db.commit()
    cur.close()
    db.close()


def _record_scan(scan_type, result=""):
    """Record that a scan just completed."""
    db = get_db()
    cur = db.cursor()
    cur.execute("""
        INSERT INTO scan_history (scan_type, started_at, finished_at, result)
        VALUES (%s, NOW(), NOW(), %s)
    """, (scan_type, result))
    db.commit()
    cur.close()
    db.close()


def _get_last_scan_dates():
    """Get the last run time for each scan type."""
    db = get_db()
    cur = db.cursor()
    cur.execute("""
        SELECT DISTINCT ON (scan_type) scan_type, finished_at, result
        FROM scan_history
        ORDER BY scan_type, finished_at DESC
    """)
    rows = cur.fetchall()
    cur.close()
    db.close()
    return {row["scan_type"]: {"last_run": row["finished_at"].isoformat() if row["finished_at"] else None, "result": row["result"]} for row in rows}


# --- Daily Auto-Scan Scheduler ---

import threading
import time as _time
from zoneinfo import ZoneInfo

_daily_scan_state = {
    "enabled": False,
    "last_run": None,
    "next_run": None,
    "status": "idle",  # idle, running_web, running_ch, running_tr1, done
    "last_result": None,
}
_daily_lock = threading.Lock()


def _get_next_8am_bst():
    """Calculate the next 8:00 AM BST/GMT (Europe/London)."""
    tz = ZoneInfo("Europe/London")
    now = datetime.now(tz)
    target = now.replace(hour=8, minute=0, second=0, microsecond=0)
    if now >= target:
        target += timedelta(days=1)
    return target


def _run_daily_scans():
    """Run all three scans sequentially."""
    from scanner import run_scan, get_scan_status
    from ch_scanner import run_ch_scan, get_ch_scan_status
    from tr1_scanner import run_tr1_scan, get_tr1_scan_status

    results = {}

    # 1. Web scan
    with _daily_lock:
        _daily_scan_state["status"] = "running_web"
    print("[daily] Starting web scan...")
    run_scan()
    while get_scan_status()["running"]:
        _time.sleep(5)
    web_status = get_scan_status()
    results["web"] = web_status.get("phase_detail", "")
    _record_scan("web", results["web"])
    print(f"[daily] Web scan done: {results['web']}")
    _time.sleep(5)

    # 2. Companies House scan
    with _daily_lock:
        _daily_scan_state["status"] = "running_ch"
    print("[daily] Starting Companies House scan (recent)...")
    run_ch_scan(recent_only=True)
    while get_ch_scan_status()["running"]:
        _time.sleep(5)
    ch_status = get_ch_scan_status()
    results["ch"] = ch_status.get("phase_detail", "")
    _record_scan("ch", results["ch"])
    print(f"[daily] CH scan done: {results['ch']}")
    _time.sleep(5)

    # 3. TR1 Direct (5,000 Investegate IDs)
    from tr1_direct import run_tr1_direct, get_direct_status
    with _daily_lock:
        _daily_scan_state["status"] = "running_tr1"
    print("[daily] Starting TR1 direct scan (recent)...")
    run_tr1_direct(recent_only=True)
    while get_direct_status()["running"]:
        _time.sleep(5)
    direct_status = get_direct_status()
    results["tr1"] = direct_status.get("phase_detail", "")
    _record_scan("tr1", results["tr1"])
    print(f"[daily] TR1 direct done: {results['tr1']}")

    return results


def _daily_scheduler_loop():
    """Background thread that checks every 60s if it's time to run."""
    while True:
        try:
            with _daily_lock:
                enabled = _daily_scan_state["enabled"]

            if enabled:
                tz = ZoneInfo("Europe/London")
                now = datetime.now(tz)
                next_run = _get_next_8am_bst()

                with _daily_lock:
                    _daily_scan_state["next_run"] = next_run.isoformat()

                # Check if we should run now (within 2 minutes of target)
                target_today = now.replace(hour=8, minute=0, second=0, microsecond=0)
                diff = (now - target_today).total_seconds()
                last_run = _daily_scan_state.get("last_run", "")
                already_ran_today = last_run and last_run[:10] == now.strftime("%Y-%m-%d")

                if 0 <= diff < 120 and not already_ran_today:
                    print(f"[daily] Triggering daily scan at {now.isoformat()}")
                    with _daily_lock:
                        _daily_scan_state["status"] = "running_web"
                        _daily_scan_state["last_run"] = now.isoformat()

                    try:
                        results = _run_daily_scans()
                        with _daily_lock:
                            _daily_scan_state["status"] = "done"
                            _daily_scan_state["last_result"] = results
                            _daily_scan_state["next_run"] = _get_next_8am_bst().isoformat()
                        print(f"[daily] All scans complete: {results}")
                    except Exception as e:
                        print(f"[daily] Scan error: {e}")
                        with _daily_lock:
                            _daily_scan_state["status"] = f"error: {e}"

        except Exception as e:
            print(f"[daily] Scheduler error: {e}")

        _time.sleep(60)  # check every minute


# Start the scheduler thread
_scheduler_thread = threading.Thread(target=_daily_scheduler_loop, daemon=True)
_scheduler_thread.start()
print("[startup] Daily scan scheduler thread started")


# --- App setup ---

print(f"[startup] DATABASE_URL set: {'yes' if os.environ.get('DATABASE_URL') else 'NO'}")

db = get_db()
init_db(db)
seed_db(db)
_init_scan_history()

# Log the final count after seeding
_cur = db.cursor()
_cur.execute("SELECT COUNT(*) as c FROM investors")
startup_count = _cur.fetchone()["c"]
_cur.close()
print(f"[startup] Investor count after init: {startup_count}")


@asynccontextmanager
async def lifespan(app):
    yield
    try:
        db.close()
    except Exception:
        pass


app = FastAPI(lifespan=lifespan)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])


# --- Models ---

class InvestorCreate(BaseModel):
    name: str
    role: Optional[str] = None
    company: Optional[str] = None
    eis_company: Optional[str] = None
    sector: Optional[str] = None
    amount: Optional[str] = None
    source_url: Optional[str] = None
    source_type: Optional[str] = None
    source_name: Optional[str] = None
    context_quote: Optional[str] = None
    linkedin_url: Optional[str] = None
    date_found: Optional[str] = None


class BatchInvestors(BaseModel):
    investors: list[InvestorCreate]


# --- Helpers ---

def row_to_dict(row):
    if row is None:
        return None
    return dict(row)


def rows_to_list(rows):
    return [dict(r) for r in rows]


# --- Endpoints ---

@app.get("/api/investors")
def list_investors(
    search: Optional[str] = Query(None),
    origin: Optional[str] = Query(None),
    entity_type: Optional[str] = Query(None),
    holding: Optional[str] = Query(None),
    source_type: Optional[str] = Query(None),
    source_name: Optional[str] = Query(None),
    sector: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=100),
    sort_by: str = Query("date_found"),
    sort_dir: str = Query("desc"),
):
    conditions = []
    params = []

    if search:
        conditions.append("(name ILIKE %s OR company ILIKE %s OR eis_company ILIKE %s OR role ILIKE %s)")
        s = f"%{search}%"
        params.extend([s, s, s, s])

    if origin == "ch":
        conditions.append("source_name = 'Companies House'")
    elif origin == "tr1":
        conditions.append("source_name = 'LSE TR1 Filing'")
    elif origin == "import":
        conditions.append("source_name = 'Excel Import'")
    elif origin == "web":
        conditions.append("(source_name IS NULL OR (source_name != 'Companies House' AND source_name != 'LSE TR1 Filing' AND source_name != 'Excel Import'))")


    if entity_type == "individual":
        conditions.append("(company != 'Organisation' OR company IS NULL)")
    elif entity_type == "organisation":
        conditions.append("company = 'Organisation'")

    if holding == "under50":
        conditions.append("""
            (context_quote ILIKE '%%25-to-50%%'
             OR (role ILIKE '%%PSC%%' 
                 AND context_quote NOT ILIKE '%%50-to-75%%'
                 AND context_quote NOT ILIKE '%%75-to-100%%'
                 AND context_quote NOT ILIKE '%%over-75%%'))
        """)
    elif holding == "50plus":
        conditions.append("""
            (context_quote ILIKE '%%50-to-75%%'
             OR context_quote ILIKE '%%75-to-100%%'
             OR context_quote ILIKE '%%over-75%%')
        """)


    if source_type:
        conditions.append("source_type = %s")
        params.append(source_type)

    if source_name:
        conditions.append("source_name = %s")
        params.append(source_name)

    if sector:
        conditions.append("sector = %s")
        params.append(sector)

    if date_from:
        conditions.append("date_found >= %s")
        params.append(date_from)

    if date_to:
        conditions.append("date_found <= %s")
        params.append(date_to)

    where = " AND ".join(conditions) if conditions else "1=1"

    # Validate sort
    allowed_sort = {"date_found", "name", "eis_company", "sector", "amount", "created_at"}
    if sort_by not in allowed_sort:
        sort_by = "date_found"
    if sort_dir not in ("asc", "desc"):
        sort_dir = "desc"

    cur = db.cursor()

    # Count
    count_sql = f"SELECT COUNT(*) as total FROM investors WHERE {where}"
    cur.execute(count_sql, params)
    total = cur.fetchone()["total"]

    # Fetch
    offset = (page - 1) * per_page
    data_sql = f"SELECT * FROM investors WHERE {where} ORDER BY {sort_by} {sort_dir} LIMIT %s OFFSET %s"
    cur.execute(data_sql, params + [per_page, offset])
    rows = cur.fetchall()
    cur.close()

    return {
        "investors": rows_to_list(rows),
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": math.ceil(total / per_page) if total > 0 else 1,
    }


@app.get("/api/investors/{investor_id}")
def get_investor(investor_id: int):
    cur = db.cursor()
    cur.execute("SELECT * FROM investors WHERE id = %s", [investor_id])
    row = cur.fetchone()
    cur.close()
    if not row:
        raise HTTPException(status_code=404, detail="Investor not found")
    return row_to_dict(row)


@app.get("/api/stats")
def get_stats():
    cur = db.cursor()

    cur.execute("SELECT COUNT(*) as c FROM investors")
    total = cur.fetchone()["c"]

    # New this week (last 7 days)
    week_ago = (datetime.now() - timedelta(days=7)).strftime("%Y-%m-%d")
    cur.execute(
        "SELECT COUNT(*) as c FROM investors WHERE date_found >= %s", [week_ago]
    )
    new_this_week = cur.fetchone()["c"]

    # Top sector
    cur.execute(
        "SELECT sector, COUNT(*) as c FROM investors GROUP BY sector ORDER BY c DESC LIMIT 1"
    )
    top_sector_row = cur.fetchone()
    top_sector = top_sector_row["sector"] if top_sector_row else "N/A"

    # Source types scanned
    cur.execute("SELECT COUNT(DISTINCT source_type) as c FROM investors")
    sources = cur.fetchone()["c"]

    # Unique sectors list
    cur.execute("SELECT DISTINCT sector FROM investors ORDER BY sector")
    sectors = [r["sector"] for r in cur.fetchall()]

    # Source types list
    cur.execute("SELECT DISTINCT source_type FROM investors ORDER BY source_type")
    source_types = [r["source_type"] for r in cur.fetchall()]

    # Source names list
    cur.execute(
        "SELECT DISTINCT source_name FROM investors WHERE source_name IS NOT NULL AND source_name != '' ORDER BY source_name"
    )
    source_names = [r["source_name"] for r in cur.fetchall()]

    # LinkedIn count
    cur.execute("SELECT COUNT(*) as c FROM investors WHERE linkedin_url IS NOT NULL AND linkedin_url != ''")
    linkedin_count = cur.fetchone()["c"]

    # Individuals vs Organisations
    cur.execute("SELECT COUNT(*) as c FROM investors WHERE company = 'Organisation'")
    org_count = cur.fetchone()["c"]
    individual_count = total - org_count

    cur.close()

    return {
        "total_investors": total,
        "new_this_week": new_this_week,
        "top_sector": top_sector,
        "sources_scanned": sources,
        "linkedin_count": linkedin_count,
        "individual_count": individual_count,
        "org_count": org_count,
        "sectors": sectors,
        "source_types": source_types,
        "source_names": source_names,
    }


def _watch_and_record(scan_type, status_fn, poll_interval=5):
    """Background thread that waits for a scan to finish and records it."""
    def _watch():
        _time.sleep(3)  # let scan start
        while True:
            s = status_fn()
            if not s.get("running", False):
                result = s.get("phase_detail", "")
                if s.get("phase") in ("done", "error"):
                    _record_scan(scan_type, result)
                break
            _time.sleep(poll_interval)
    threading.Thread(target=_watch, daemon=True).start()


@app.post("/api/scan")
def trigger_scan():
    from scanner import run_scan, get_scan_status
    status = get_scan_status()
    if status["running"]:
        return {"status": "already_running", "message": "A scan is already in progress."}
    started = run_scan()
    if started:
        _watch_and_record("web", get_scan_status)
        return {"status": "started", "message": "Scan started. Poll /api/scan/status for progress."}
    return {"status": "error", "message": "Failed to start scan."}


@app.get("/api/scan/status")
def scan_status():
    from scanner import get_scan_status
    return get_scan_status()


@app.post("/api/ch-scan")
def trigger_ch_scan(recent_only: bool = Query(False)):
    from ch_scanner import run_ch_scan, get_ch_scan_status
    status = get_ch_scan_status()
    if status["running"]:
        return {"status": "already_running", "message": "A Companies House scan is already in progress."}
    started = run_ch_scan(recent_only=recent_only)
    mode = "recent" if recent_only else "full"
    if started:
        _watch_and_record("ch", get_ch_scan_status)
        return {"status": "started", "message": f"Companies House {mode} scan started."}
    return {"status": "error", "message": "Failed to start Companies House scan."}


@app.get("/api/ch-scan/status")
def ch_scan_status():
    from ch_scanner import get_ch_scan_status
    return get_ch_scan_status()


@app.post("/api/tr1-scan")
def trigger_tr1_scan(days_back: int = Query(30, ge=1, le=3650)):
    from tr1_scanner import run_tr1_scan, get_tr1_scan_status
    status = get_tr1_scan_status()
    if status["running"]:
        return {"status": "already_running", "message": "A TR1 scan is already in progress."}
    started = run_tr1_scan(days_back=days_back)
    if started:
        return {"status": "started", "message": f"TR1 scan started ({days_back} days). Poll /api/tr1-scan/status for progress."}
    return {"status": "error", "message": "Failed to start TR1 scan."}


@app.get("/api/tr1-scan/status")
def tr1_scan_status():
    from tr1_scanner import get_tr1_scan_status
    return get_tr1_scan_status()


# --- TR1 Sweep Endpoints ---

@app.post("/api/tr1-sweep")
def trigger_tr1_sweep():
    from tr1_sweep import run_tr1_sweep, get_sweep_status
    status = get_sweep_status()
    if status["running"]:
        return {"status": "already_running", "message": "TR1 sweep already in progress."}
    started = run_tr1_sweep()
    if started:
        return {"status": "started", "message": "TR1 company sweep started."}
    return {"status": "error", "message": "Failed to start TR1 sweep."}


@app.get("/api/tr1-sweep/status")
def tr1_sweep_status():
    from tr1_sweep import get_sweep_status
    return get_sweep_status()


# --- TR1 Direct Endpoints ---

@app.post("/api/tr1-direct")
def trigger_tr1_direct(max_market_cap: int = Query(0, ge=0), recent_only: bool = Query(False)):
    from tr1_direct import run_tr1_direct, get_direct_status
    status = get_direct_status()
    if status["running"]:
        return {"status": "already_running", "message": "TR1 direct scan already in progress."}
    started = run_tr1_direct(max_market_cap=max_market_cap, recent_only=recent_only)
    cap_msg = f" (under \u00a3{max_market_cap:,}M)" if max_market_cap > 0 else ""
    mode = "recent" if recent_only else "full"
    if started:
        _watch_and_record("tr1", get_direct_status)
        return {"status": "started", "message": f"TR1 {mode} scan started{cap_msg}."}
    return {"status": "error", "message": "Failed to start TR1 direct scan."}


@app.post("/api/tr1-direct/stop")
def stop_tr1_direct():
    from tr1_direct import stop_direct_scan
    stopped = stop_direct_scan()
    if stopped:
        return {"status": "stopping", "message": "Stop requested. Will stop after current batch."}
    return {"status": "not_running", "message": "No scan is running."}


@app.get("/api/tr1-direct/status")
def tr1_direct_status():
    from tr1_direct import get_direct_status
    return get_direct_status()


# --- Investor Research Endpoint ---

@app.get("/api/investor/{investor_id}/research")
def research_investor(investor_id: int):
    """Research an investor and return a summary profile."""
    import httpx as _httpx

    cur = db.cursor()
    cur.execute("SELECT * FROM investors WHERE id = %s", (investor_id,))
    inv = cur.fetchone()
    cur.close()

    if not inv:
        raise HTTPException(status_code=404, detail="Investor not found")

    name = inv["name"] or ""
    company = inv.get("eis_company") or ""
    role = inv.get("role") or ""
    linkedin = inv.get("linkedin_url") or ""
    source = inv.get("source_name") or ""
    context = (inv.get("context_quote") or "")[:300]  # Truncate long context

    # Build research prompt
    prompt = f"""Research this UK investor and provide a concise professional summary.

Name: {name}
Role: {role}
Company/Issuer: {company}
Source: {source}
Context: {context}
LinkedIn: {linkedin or 'Not available'}

Provide a summary covering:
1. **Background**: Who they are, their professional background
2. **Investment activity**: Known investments, sectors they invest in, investment style
3. **Current role**: Current company and position
4. **Notable**: Any notable achievements, board positions, or public profile

If LinkedIn URL is provided, reference their likely professional history based on the URL handle.
Be factual and concise. If information is limited, say so rather than speculate.
Format as plain text paragraphs, not markdown."""

    gemini_key = os.environ.get("GEMINI_API_KEY", "")
    if not gemini_key:
        return {"summary": "Gemini API key not configured. Cannot generate research summary."}

    try:
        resp = _httpx.post(
            "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent",
            headers={"x-goog-api-key": gemini_key, "Content-Type": "application/json"},
            json={
                "contents": [{"parts": [{"text": prompt}]}],
                "generationConfig": {"temperature": 0.3, "maxOutputTokens": 1000},
            },
            timeout=30,
        )
        if resp.status_code == 200:
            data = resp.json()
            text = data.get("candidates", [{}])[0].get("content", {}).get("parts", [{}])[0].get("text", "")
            # Save summary to bio field
            if text:
                cur2 = db.cursor()
                cur2.execute(
                    "UPDATE investors SET bio = %s WHERE id = %s",
                    (text.strip(), investor_id)
                )
                db.commit()
                cur2.close()
            return {"summary": text.strip()}
        else:
            err_body = resp.text[:200] if resp.text else "no details"
            print(f"[research] Gemini {resp.status_code}: {err_body}")
            return {"summary": f"Gemini returned {resp.status_code}: {err_body}"}
    except Exception as e:
        return {"summary": f"Research failed: {str(e)}"}


# --- LinkedIn Enrichment Endpoints ---

@app.post("/api/enrich")
def trigger_enrichment():
    from enricher import run_enrichment, get_enrich_status
    status = get_enrich_status()
    if status["running"]:
        return {"status": "already_running", "message": "Enrichment already in progress."}
    started = run_enrichment()
    if started:
        return {"status": "started", "message": "LinkedIn enrichment started."}
    return {"status": "error", "message": "Failed to start enrichment."}


@app.post("/api/enrich/stop")
def stop_enrichment():
    from enricher import stop_enrichment as _stop
    stopped = _stop()
    if stopped:
        return {"status": "stopping", "message": "Stop requested."}
    return {"status": "not_running"}


@app.get("/api/enrich/status")
def enrich_status():
    from enricher import get_enrich_status
    return get_enrich_status()


# --- Scan History Endpoint ---

@app.get("/api/scan-history")
def scan_history():
    return _get_last_scan_dates()


# --- Daily Update Endpoints ---

@app.get("/api/daily/status")
def daily_status():
    with _daily_lock:
        return dict(_daily_scan_state)


@app.post("/api/daily/enable")
def daily_enable():
    with _daily_lock:
        _daily_scan_state["enabled"] = True
        _daily_scan_state["next_run"] = _get_next_8am_bst().isoformat()
    print("[daily] Auto-scan ENABLED")
    return {"status": "enabled", "next_run": _daily_scan_state["next_run"]}


@app.post("/api/daily/disable")
def daily_disable():
    with _daily_lock:
        _daily_scan_state["enabled"] = False
        _daily_scan_state["next_run"] = None
    print("[daily] Auto-scan DISABLED")
    return {"status": "disabled"}


# Keep legacy endpoint for backward compatibility
@app.post("/api/collect")
def trigger_collection():
    from scanner import run_scan, get_scan_status
    status = get_scan_status()
    if status["running"]:
        return {"status": "already_running", "message": "A scan is already in progress."}
    run_scan()
    return {"status": "started", "message": "Scan started."}


@app.post("/api/investors/batch", status_code=201)
def batch_upsert(batch: BatchInvestors):
    inserted = 0
    skipped = 0
    cur = db.cursor()

    for inv in batch.investors:
        # Check for duplicate by name + eis_company
        cur.execute(
            "SELECT id FROM investors WHERE name = %s AND eis_company = %s",
            [inv.name, inv.eis_company]
        )
        existing = cur.fetchone()

        if existing:
            # Update existing record
            cur.execute("""
                UPDATE investors SET role=%s, company=%s, sector=%s, amount=%s,
                source_url=%s, source_type=%s, source_name=%s, context_quote=%s,
                linkedin_url=%s, date_found=%s
                WHERE id=%s
            """, (
                inv.role, inv.company, inv.sector, inv.amount,
                inv.source_url, inv.source_type, inv.source_name, inv.context_quote,
                inv.linkedin_url, inv.date_found, existing["id"]
            ))
            skipped += 1
        else:
            cur.execute("""
                INSERT INTO investors (name, role, company, eis_company, sector, amount,
                source_url, source_type, source_name, context_quote, linkedin_url, date_found)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                inv.name, inv.role, inv.company, inv.eis_company, inv.sector, inv.amount,
                inv.source_url, inv.source_type, inv.source_name, inv.context_quote,
                inv.linkedin_url, inv.date_found
            ))
            inserted += 1

    db.commit()
    cur.close()
    return {"inserted": inserted, "updated": skipped, "total": len(batch.investors)}


@app.get("/api/export/last")
def get_last_export():
    """Return info about the last 'new' export."""
    cur = db.cursor()
    cur.execute(
        "SELECT exported_at, investor_count FROM export_log WHERE export_type = 'new' ORDER BY id DESC LIMIT 1"
    )
    row = cur.fetchone()
    if row:
        # Count investors added since that export
        cur.execute(
            "SELECT COUNT(*) as c FROM investors WHERE created_at > %s", [row["exported_at"]]
        )
        new_since = cur.fetchone()["c"]
        cur.close()
        return {
            "last_exported_at": row["exported_at"],
            "last_export_count": row["investor_count"],
            "new_since_last_export": new_since,
        }
    else:
        cur.execute("SELECT COUNT(*) as c FROM investors")
        total = cur.fetchone()["c"]
        cur.close()
        return {
            "last_exported_at": None,
            "last_export_count": 0,
            "new_since_last_export": total,
        }


def build_excel(investors, title_text, subtitle_text):
    """Build a formatted Excel workbook from a list of investor dicts."""
    wb = Workbook()
    ws = wb.active
    ws.title = "EIS Investors"

    # ── Colours ──
    HEADER_FILL = PatternFill(start_color="1B3A4B", end_color="1B3A4B", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    EVEN_FILL = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
    ODD_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    BODY_FONT = Font(name="Calibri", size=10, color="1A1A1A")
    LINK_FONT = Font(name="Calibri", size=10, color="2980B9", underline="single")
    MUTED_FONT = Font(name="Calibri", size=10, color="808080", italic=True)
    AMOUNT_FONT = Font(name="Calibri", size=10, bold=True, color="1A6B3C")
    THIN_BORDER = Border(bottom=Side(style="thin", color="E0E0E0"))
    HEADER_BORDER = Border(bottom=Side(style="medium", color="0F2B3A"))

    # ── Title row ──
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = title_text
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="1B3A4B")
    title_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 36

    # ── Subtitle row ──
    ws.merge_cells("A2:K2")
    sub_cell = ws["A2"]
    sub_cell.value = subtitle_text
    sub_cell.font = Font(name="Calibri", size=10, color="666666")
    sub_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 22

    # ── Spacer row ──
    ws.row_dimensions[3].height = 8

    # ── Headers (row 4) ──
    headers = [
        ("Name", 26),
        ("Role", 24),
        ("Company", 24),
        ("EIS Company", 28),
        ("Sector", 26),
        ("Amount", 16),
        ("Source", 20),
        ("Source Type", 14),
        ("Date Found", 14),
        ("LinkedIn", 32),
        ("Source URL", 40),
    ]

    header_row = 4
    for col_idx, (header_name, col_width) in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = HEADER_BORDER
        ws.column_dimensions[cell.column_letter].width = col_width

    ws.row_dimensions[header_row].height = 30

    # ── Data rows ──
    for row_idx, inv in enumerate(investors, header_row + 1):
        fill = EVEN_FILL if (row_idx - header_row) % 2 == 0 else ODD_FILL

        values = [
            inv.get("name", ""),
            inv.get("role", ""),
            inv.get("company", ""),
            inv.get("eis_company", ""),
            inv.get("sector", ""),
            inv.get("amount", ""),
            inv.get("source_name", ""),
            inv.get("source_type", ""),
            inv.get("date_found", ""),
            inv.get("linkedin_url", "") or "",
            inv.get("source_url", "") or "",
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)

            if col_idx == 1:
                cell.font = Font(name="Calibri", size=10, bold=True, color="1A1A1A")
            elif col_idx == 6:
                if val and val not in ("Undisclosed", "undisclosed", "Not disclosed", "not disclosed", ""):
                    cell.font = AMOUNT_FONT
                else:
                    cell.font = MUTED_FONT
                    cell.value = "Undisclosed"
            elif col_idx == 10 and val:
                cell.font = LINK_FONT
                cell.hyperlink = val
                cell.value = val
            elif col_idx == 11 and val:
                cell.font = LINK_FONT
                cell.hyperlink = val
                cell.value = val
            else:
                cell.font = BODY_FONT

        ws.row_dimensions[row_idx].height = 24

    # ── Freeze panes ──
    ws.freeze_panes = f"A{header_row + 1}"

    # ── Auto-filter ──
    last_row = header_row + len(investors)
    last_col_letter = ws.cell(row=header_row, column=len(headers)).column_letter
    ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{last_row}"

    # ── Print settings ──
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@app.get("/api/export/excel")
def export_excel(
    search: Optional[str] = Query(None),
    source_type: Optional[str] = Query(None),
    sector: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    sort_by: str = Query("date_found"),
    sort_dir: str = Query("desc"),
):
    """Export all matching investors as a formatted Excel workbook."""
    conditions = []
    params = []

    if search:
        conditions.append("(name ILIKE %s OR company ILIKE %s OR eis_company ILIKE %s OR role ILIKE %s)")
        s = f"%{search}%"
        params.extend([s, s, s, s])
    if source_type:
        conditions.append("source_type = %s")
        params.append(source_type)
    if sector:
        conditions.append("sector = %s")
        params.append(sector)
    if date_from:
        conditions.append("date_found >= %s")
        params.append(date_from)
    if date_to:
        conditions.append("date_found <= %s")
        params.append(date_to)

    where = " AND ".join(conditions) if conditions else "1=1"
    allowed_sort = {"date_found", "name", "eis_company", "sector", "amount", "created_at"}
    if sort_by not in allowed_sort:
        sort_by = "date_found"
    if sort_dir not in ("asc", "desc"):
        sort_dir = "desc"

    cur = db.cursor()
    cur.execute(
        f"SELECT * FROM investors WHERE {where} ORDER BY {sort_by} {sort_dir}", params
    )
    rows = cur.fetchall()
    cur.close()
    investors = [dict(r) for r in rows]

    now_str = datetime.now().strftime('%d %B %Y at %H:%M')
    buffer = build_excel(
        investors,
        title_text="EIS Investor Collector \u2014 Full Export",
        subtitle_text=f"Generated {now_str}  \u00b7  {len(investors)} investors",
    )

    filename = f"eis_investors_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/export/excel-new")
def export_excel_new():
    """Export only investors added since the last 'new' export, then record this export."""
    cur = db.cursor()

    # Find last export timestamp
    cur.execute(
        "SELECT exported_at FROM export_log WHERE export_type = 'new' ORDER BY id DESC LIMIT 1"
    )
    last_row = cur.fetchone()
    last_exported_at = last_row["exported_at"] if last_row else None

    if last_exported_at:
        cur.execute(
            "SELECT * FROM investors WHERE created_at > %s ORDER BY created_at DESC",
            [last_exported_at],
        )
    else:
        cur.execute("SELECT * FROM investors ORDER BY created_at DESC")

    rows = cur.fetchall()
    investors = [dict(r) for r in rows]

    if len(investors) == 0:
        cur.close()
        raise HTTPException(status_code=404, detail="No new investors since last export.")

    now = datetime.now()
    now_str = now.strftime('%d %B %Y at %H:%M')
    since_str = last_exported_at.strftime("%Y-%m-%d %H:%M") if last_exported_at else "the beginning"
    buffer = build_excel(
        investors,
        title_text="EIS Investor Collector \u2014 New Since Last Export",
        subtitle_text=f"Generated {now_str}  \u00b7  {len(investors)} new investors since {since_str}",
    )

    # Record this export
    cur.execute(
        "INSERT INTO export_log (exported_at, investor_count, export_type) VALUES (%s, %s, 'new')",
        [now, len(investors)],
    )
    db.commit()
    cur.close()

    filename = f"eis_investors_new_{now.strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/email/new-investors")
def get_new_investors_for_email():
    """Return investors added since last email, and record the email event."""
    cur = db.cursor()
    cur.execute(
        "SELECT emailed_at FROM email_log ORDER BY id DESC LIMIT 1"
    )
    last_row = cur.fetchone()
    last_emailed_at = last_row["emailed_at"] if last_row else None

    if last_emailed_at:
        cur.execute(
            "SELECT name, role, company, eis_company, sector, amount, source_name, date_found "
            "FROM investors WHERE created_at > %s ORDER BY created_at DESC",
            [last_emailed_at],
        )
    else:
        cur.execute(
            "SELECT name, role, company, eis_company, sector, amount, source_name, date_found "
            "FROM investors ORDER BY created_at DESC"
        )

    rows = cur.fetchall()
    cur.close()
    investors = [dict(r) for r in rows]
    return {
        "investors": investors,
        "count": len(investors),
        "since": last_emailed_at,
    }


@app.post("/api/email/mark-sent")
def mark_email_sent(count: int = Query(0)):
    """Record that an email digest was sent."""
    now = datetime.now()
    cur = db.cursor()
    cur.execute(
        "INSERT INTO email_log (emailed_at, investor_count) VALUES (%s, %s)",
        [now, count],
    )
    db.commit()
    cur.close()
    return {"status": "recorded", "emailed_at": now.strftime("%Y-%m-%dT%H:%M:%S"), "count": count}


# --- Import from Excel ---

@app.post("/api/import/excel")
async def import_excel(file: UploadFile = File(...)):
    """Import contacts from an uploaded Excel file.
    
    Reads the Excel, maps columns intelligently to our schema,
    deduplicates against existing records, and inserts new ones.
    """
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx or .xls file.")

    try:
        contents = await file.read()
        wb = load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active

        # Read all rows
        all_rows_raw = list(ws.iter_rows(values_only=True))
        if not all_rows_raw:
            return {"status": "empty", "message": "File is empty.", "imported": 0, "duplicates": 0}

        # Find the header row — scan first 5 rows for one with 3+ cells
        # and a name-like column
        header_row_idx = 0
        name_keywords = ["name", "investor", "contact", "person"]

        for row_idx in range(min(5, len(all_rows_raw))):
            cells = [str(v).strip().lower() if v else "" for v in all_rows_raw[row_idx]]
            non_empty = sum(1 for c in cells if c)
            has_name = any(any(kw == c or kw in c for kw in name_keywords) for c in cells if c)
            if non_empty >= 3 and has_name:
                header_row_idx = row_idx
                break

        # Build clean header list
        raw_header_row = all_rows_raw[header_row_idx]
        headers = []
        for i, v in enumerate(raw_header_row):
            h = str(v).strip().lower() if v else f"col_{i}"
            # Remove non-breaking spaces and other invisible chars
            h = h.replace('\u00a0', ' ').replace('\u200b', '').strip()
            headers.append(h)

        # Map columns
        col_map = _map_columns(headers)

        # If mapping still failed, just assign columns by position
        # based on what the app's own export produces
        if "name" not in col_map and len(headers) >= 6:
            # Assume standard order: name, role, company, eis_company, sector, amount, ...
            positional = ["name", "role", "company", "eis_company", "sector", "amount",
                          "source_name", "source_type", "date_found", "linkedin_url", "source_url"]
            col_map = {}
            for idx, field in enumerate(positional):
                if idx < len(headers):
                    col_map[field] = idx

        if "name" not in col_map:
            raise HTTPException(
                status_code=400,
                detail=f"Could not find a 'name' column. Headers: {headers}. Mapped: {col_map}"
            )

        # Read all data rows (everything after the header row)
        rows_data = []
        for row in all_rows_raw[header_row_idx + 1:]:
            if not any(row):  # skip empty rows
                continue
            investor = {}
            for field, col_idx in col_map.items():
                val = row[col_idx] if col_idx < len(row) else None
                investor[field] = str(val).strip() if val is not None else ""
            rows_data.append(investor)

        wb.close()

        if not rows_data:
            return {"status": "empty", "message": "No data rows found in the file.", "imported": 0, "duplicates": 0}

        # Insert into database with dedup
        today = datetime.now().strftime("%Y-%m-%d")
        idb = get_db()
        cur = idb.cursor()
        inserted = 0
        duplicated = 0

        for inv in rows_data:
            name = inv.get("name", "").strip()
            if not name:
                continue

            eis_company = inv.get("eis_company", "").strip()

            # Check for duplicate
            cur.execute(
                "SELECT id FROM investors WHERE name = %s AND eis_company = %s",
                (name, eis_company)
            )
            if cur.fetchone():
                duplicated += 1
                continue

            cur.execute("""
                INSERT INTO investors (name, role, company, eis_company, sector, amount,
                source_url, source_type, source_name, context_quote, linkedin_url, date_found)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                name,
                inv.get("role", ""),
                inv.get("company", ""),
                eis_company,
                inv.get("sector", ""),
                inv.get("amount", ""),
                inv.get("source_url", ""),
                inv.get("source_type", "Import"),
                inv.get("source_name", "Excel Import"),
                inv.get("context_quote", f"Imported from {file.filename}"),
                inv.get("linkedin_url", ""),
                inv.get("date_found", today),
            ))
            inserted += 1

        idb.commit()
        cur.close()
        idb.close()

        return {
            "status": "ok",
            "message": f"Imported {inserted} new contacts, {duplicated} duplicates skipped.",
            "imported": inserted,
            "duplicates": duplicated,
            "total_rows": len(rows_data),
            "columns_mapped": {k: headers[v] for k, v in col_map.items()},
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")


def _map_columns(headers):
    """Map Excel column headers to our investor schema fields.
    
    Strategy: first pass tries exact matches, second pass tries substring.
    Each column can only be assigned to one field.
    """
    col_map = {}
    used_cols = set()
    
    # Exact matches first (header == keyword)
    exact = {
        "name": ["name", "investor name", "full name", "contact name", "person"],
        "role": ["role", "title", "position", "job title"],
        "company": ["company", "firm", "organisation", "organization", "employer"],
        "eis_company": ["eis company", "eis_company", "investee", "portfolio company", "target company", "invested in"],
        "sector": ["sector", "industry", "vertical", "category"],
        "amount": ["amount", "investment amount", "invested"],
        "source_url": ["source url", "url", "link", "website"],
        "source_type": ["source type", "type"],
        "source_name": ["source", "source name", "data source", "origin"],
        "linkedin_url": ["linkedin", "linkedin url", "profile", "linkedin profile"],
        "context_quote": ["context", "notes", "quote", "description", "details", "comments"],
        "date_found": ["date found", "date", "date_found", "found date"],
    }
    
    # Pass 1: exact match (header equals keyword)
    for field, keywords in exact.items():
        for i, header in enumerate(headers):
            if i in used_cols:
                continue
            if header in keywords:
                col_map[field] = i
                used_cols.add(i)
                break
    
    # Pass 2: substring match for any fields still unmapped
    substring = {
        "name": ["name", "investor", "contact", "person"],
        "role": ["role", "title", "position"],
        "company": ["company", "firm", "organisation"],
        "eis_company": ["eis", "investee", "portfolio", "target"],
        "sector": ["sector", "industry"],
        "amount": ["amount", "invested", "size", "value"],
        "source_url": ["url", "link"],
        "source_name": ["source"],
        "linkedin_url": ["linkedin"],
        "context_quote": ["note", "context", "comment", "detail"],
        "date_found": ["date"],
    }
    
    for field, keywords in substring.items():
        if field in col_map:
            continue
        for i, header in enumerate(headers):
            if i in used_cols:
                continue
            if any(kw in header for kw in keywords):
                col_map[field] = i
                used_cols.add(i)
                break
    
    return col_map


# --- Serve static frontend ---
STATIC_DIR = Path(__file__).parent / "static"
if STATIC_DIR.is_dir():
    @app.get("/")
    async def serve_index():
        return FileResponse(STATIC_DIR / "index.html")

    app.mount("/", StaticFiles(directory=str(STATIC_DIR)), name="static")


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
